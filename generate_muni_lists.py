# -*- coding: utf-8 -*-
"""総務省「全国地方公共団体コード」(muni_master.xlsx)から
全47都道府県分のmuni_lists.py用データを生成する使い捨てスクリプト。
既存の7都県(tokyo/kanagawa/ibaraki/tochigi/gunma/saitama/chiba)は手作業データを維持し、
残り40道府県分のみを生成してJSONに書き出す。
"""
import json
import re
import openpyxl

XLSX_PATH = "muni_master.xlsx"

# 都道府県名(漢字) -> (romajiキー, 2桁コード)
PREF_KEY_MAP = {
    "北海道": ("hokkaido", "01"), "青森県": ("aomori", "02"), "岩手県": ("iwate", "03"),
    "宮城県": ("miyagi", "04"), "秋田県": ("akita", "05"), "山形県": ("yamagata", "06"),
    "福島県": ("fukushima", "07"), "茨城県": ("ibaraki", "08"), "栃木県": ("tochigi", "09"),
    "群馬県": ("gunma", "10"), "埼玉県": ("saitama", "11"), "千葉県": ("chiba", "12"),
    "東京都": ("tokyo", "13"), "神奈川県": ("kanagawa", "14"), "新潟県": ("niigata", "15"),
    "富山県": ("toyama", "16"), "石川県": ("ishikawa", "17"), "福井県": ("fukui", "18"),
    "山梨県": ("yamanashi", "19"), "長野県": ("nagano", "20"), "岐阜県": ("gifu", "21"),
    "静岡県": ("shizuoka", "22"), "愛知県": ("aichi", "23"), "三重県": ("mie", "24"),
    "滋賀県": ("shiga", "25"), "京都府": ("kyoto", "26"), "大阪府": ("osaka", "27"),
    "兵庫県": ("hyogo", "28"), "奈良県": ("nara", "29"), "和歌山県": ("wakayama", "30"),
    "鳥取県": ("tottori", "31"), "島根県": ("shimane", "32"), "岡山県": ("okayama", "33"),
    "広島県": ("hiroshima", "34"), "山口県": ("yamaguchi", "35"), "徳島県": ("tokushima", "36"),
    "香川県": ("kagawa", "37"), "愛媛県": ("ehime", "38"), "高知県": ("kochi", "39"),
    "福岡県": ("fukuoka", "40"), "佐賀県": ("saga", "41"), "長崎県": ("nagasaki", "42"),
    "熊本県": ("kumamoto", "43"), "大分県": ("oita", "44"), "宮崎県": ("miyazaki", "45"),
    "鹿児島県": ("kagoshima", "46"), "沖縄県": ("okinawa", "47"),
}

ALREADY_BUILT = {"tokyo", "kanagawa", "ibaraki", "tochigi", "gunma", "saitama", "chiba"}


def load_designated_cities():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws2 = wb[wb.sheetnames[1]]
    rows = [r for r in ws2.iter_rows(values_only=True)][1:]
    groups = []
    cur = None
    for code, prefname, name, *_ in rows:
        if name is None:
            continue
        code5 = str(code)[:5]
        if re.fullmatch(r".+市", name):
            cur = {"pref_name": prefname, "name": name, "code": code5, "wards": []}
            groups.append(cur)
        else:
            cur["wards"].append(code5)
    return groups


def load_all_municipalities():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws1 = wb[wb.sheetnames[0]]
    rows = [r for r in ws1.iter_rows(values_only=True)][1:]
    by_pref = {}
    for code, prefname, name, *_ in rows:
        if name is None:
            continue  # 都道府県そのものの行はスキップ
        code5 = str(code)[:5]
        by_pref.setdefault(prefname, []).append((name, code5))
    return by_pref


def main():
    designated = load_designated_cities()
    designated_by_pref = {}
    for g in designated:
        designated_by_pref.setdefault(g["pref_name"], []).append(g)

    all_munis = load_all_municipalities()

    result = {}
    for pref_name, (key, pref_code) in PREF_KEY_MAP.items():
        if key in ALREADY_BUILT:
            continue
        munis = all_munis.get(pref_name, [])
        dcs = designated_by_pref.get(pref_name, [])
        dc_codes = {dc["code"] for dc in dcs}
        # 政令指定都市本体(区の親)の行はmunicipalitiesから除外(designated_citiesで別途union)
        munis_filtered = [(n, c) for (n, c) in munis if c not in dc_codes]
        entry = {
            "name": pref_name,
            "pref_code": pref_code,
            "municipalities": munis_filtered,
        }
        if dcs:
            entry["designated_cities"] = [
                {"name": dc["name"], "code": dc["code"], "wards": dc["wards"]} for dc in dcs
            ]
        result[key] = entry

    with open("new_prefectures.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=1)
    print(f"生成完了: {len(result)}道府県分 -> new_prefectures.json")
    for key, e in result.items():
        dc_note = f" (指定都市{len(e.get('designated_cities', []))})" if e.get("designated_cities") else ""
        print(f"  {key}: {e['name']} 市区町村{len(e['municipalities'])}件{dc_note}")


if __name__ == "__main__":
    main()
